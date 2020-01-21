from openpyxl import load_workbook

wb = load_workbook('C:/Users/林琮達/Desktop/Dataset.xlsx')
print(wb.sheetnames)
ws = wb.active
print(ws['A1'].value)

"""
for row in ws.rows:
    for cell in row:
        print(cell.value)
"""
itemtable = []
print("start of columns")
for column in ws.columns:
    for cell in column:
         itemtable.append(cell.value)

print(itemtable[7])
print(len(itemtable))